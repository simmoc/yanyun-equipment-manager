import openpyxl
import os
import re

files = [
    "裂石钧105阶竞速轴属性毕业率进阶计算器1.7.xlsx",
    "裂石威105阶竞速轴属性毕业率进阶计算器0.5.xlsx",
    "鸣金虹105阶竞速轴属性毕业率进阶计算器0.1.xlsx",
    "鸣金影105阶竞速轴属性毕业率进阶计算器0.1.xlsx",
    "破竹尘105阶竞速轴属性毕业率进阶计算器0.3.xlsx",
    "破竹风100级竞速轴属性毕业率计算器5.-1 copy.xlsx",
    "破竹鸢105阶竞速轴属性毕业率进阶计算器1.1.xlsx", # Noticed version 1.1 in my file list above? wait.
    "破竹鸢105阶竞速轴属性毕业率进阶计算器0.7.xlsx",
    "牵丝霖105阶竞速轴属性毕业率进阶计算器0.4.xlsx",
    "牵丝玉100级竞速轴属性毕业率计算器5.xlsx"
]

# Check actual files in dir to be sure
# base_excel/破竹鸢105阶竞速轴属性毕业率进阶计算器0.7.xlsx
# List dir again if needed.

base_path = "/Users/simmoc/project/yanyun-equipment-manager/base_excel"

# Get actual list of xlsx files
actual_files = [f for f in os.listdir(base_path) if f.endswith('.xlsx')]

results = {}

for filename in actual_files:
    path = os.path.join(base_path, filename)
    print(f"Processing {filename}...")
    try:
        # Load workbook
        wb_formula = openpyxl.load_workbook(path, data_only=False)
        if '期望' in wb_formula.sheetnames:
            sheet_formula = wb_formula['期望']
        else:
            sheet_formula = wb_formula.active
            
        wb_value = openpyxl.load_workbook(path, data_only=True)
        if '期望' in wb_value.sheetnames:
            sheet_value = wb_value['期望']
        else:
            sheet_value = wb_value.active
        
        battle_time = sheet_value['I8'].value
        adps_current = sheet_value['I12'].value
        grad_rate = sheet_value['I16'].value
        grad_formula = sheet_formula['I16'].value
        
        # Extract target ADPS from formula
        target_adps = None
        if isinstance(grad_formula, str) and "/" in grad_formula:
            parts = grad_formula.split("/")
            if len(parts) > 1:
                # Formula might be like =I12/115456.78 or =I12/'RD'!J1
                match = re.search(r"[\d\.]+", parts[1])
                if match:
                    target_adps = float(match.group())
        
        if (target_adps is None or target_adps < 1000) and adps_current and grad_rate:
             target_adps = float(adps_current) / float(grad_rate)
             
        # Normalize name
        name_match = re.match(r"(鸣金|破竹|裂石|牵丝)(.)", filename)
        name = name_match.group(0) if name_match else filename
        
        results[name] = {
            "filename": filename,
            "battle_time": battle_time,
            "target_adps": target_adps,
            "adps_current": adps_current,
            "grad_rate": grad_rate
        }
        
    except Exception as e:
        print(f"Error processing {filename}: {e}")

print("\n--- Summary Results ---")
import json
print(json.dumps(results, indent=2, ensure_ascii=False))
