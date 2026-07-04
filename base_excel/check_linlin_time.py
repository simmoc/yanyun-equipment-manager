import openpyxl
import os

files = [
    "牵丝霖105阶竞速轴属性毕业率进阶计算器0.4.xlsx",
]

base_path = "/Users/simmoc/project/yanyun-equipment-manager/base_excel"

for filename in files:
    path = os.path.join(base_path, filename)
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'RD' in wb.sheetnames:
        sheet = wb['RD']
        # Check Column C (Duration?) or similar
        print(f"File: {filename}, Sheet: RD")
        for r in range(1, 10):
            row_values = [sheet.cell(row=r, column=c).value for c in range(1, 10)]
            print(f"Row {r}: {row_values}")
