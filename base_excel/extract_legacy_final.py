import openpyxl
import os
import re

files = [
    "牵丝玉100级竞速轴属性毕业率计算器5.xlsx",
    "破竹风100级竞速轴属性毕业率计算器5.-1 copy.xlsx"
]

base_path = "/Users/simmoc/project/yanyun-equipment-manager/base_excel"

for filename in files:
    path = os.path.join(base_path, filename)
    wb_formula = openpyxl.load_workbook(path, data_only=False)
    sheet_formula = wb_formula['期望']
    
    wb_value = openpyxl.load_workbook(path, data_only=True)
    sheet_value = wb_value['期望']
    
    print(f"\nFile: {filename}")
    
    # B40 usually contains =XX / [TARGET]
    grad_formula = sheet_formula['B40'].value
    print(f"B40 formula: {grad_formula}")
    
    # Search for Battle Time (秒, s, 时间)
    battle_time = None
    for r in range(1, 100):
        for c in range(1, 30):
            val = sheet_value.cell(row=r, column=c).value
            if val and isinstance(val, str) and ("战斗时间" in val or "总耗时" in val or "秒数" in val):
                 print(f"[{r},{c}] {val} = {sheet_value.cell(row=r, column=c+1).value}")
                 battle_time = sheet_value.cell(row=r, column=c+1).value
            
            if val and isinstance(val, str) and ("ADPS" in val or "秒伤" in val) and "毕业率" not in val:
                 print(f"[{r},{c}] {val} = {sheet_value.cell(row=r, column=c+1).value}")
    
    if battle_time is None:
        # Check row 8 column 9 for 破竹风 string?
        # File: 破竹风100级... Row 8 Column 9: 忘川泥鱼易水断石\n记得填真实延迟
        # This looks like it might be where time is if I look at legacy files.
        # Let's check 破竹风 combat time from RD sheet?
        pass
