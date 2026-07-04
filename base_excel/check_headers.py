import openpyxl
import os

files = [
    "鸣金影105阶竞速轴属性毕业率进阶计算器0.1.xlsx",
]

base_path = "/Users/simmoc/project/yanyun-equipment-manager/base_excel"

for filename in files:
    path = os.path.join(base_path, filename)
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb['期望']
    
    print(f"File: {filename}")
    for col in ['L', 'M', 'N', 'O', 'P', 'Q', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BL']:
        print(f"Column {col} Header (Row 1): {sheet.cell(row=1, column=openpyxl.utils.column_index_from_string(col)).value}")
        print(f"Column {col} Header (Row 24): {sheet.cell(row=24, column=openpyxl.utils.column_index_from_string(col)).value}")
