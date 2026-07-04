import openpyxl
import os

files = [
    "牵丝玉100级竞速轴属性毕业率计算器5.xlsx",
    "破竹风100级竞速轴属性毕业率计算器5.-1 copy.xlsx",
    "牵丝霖105阶竞速轴属性毕业率进阶计算器0.4.xlsx"
]

base_path = "/Users/simmoc/project/yanyun-equipment-manager/base_excel"

for filename in files:
    path = os.path.join(base_path, filename)
    wb = openpyxl.load_workbook(path, read_only=True)
    print(f"\nFile: {filename}")
    print(f"Sheets: {wb.sheetnames}")
    active = wb.active
    print(f"Active Sheet: {active.title}")
    # Inspect surrounding cells of I8, I12, I16
    sheet = wb[active.title]
    if '期望' in wb.sheetnames:
        sheet = wb['期望']
    
    print(f"I7: {sheet['I7'].value}, I8: {sheet['I8'].value}, I9: {sheet['I9'].value}")
    print(f"I11: {sheet['I11'].value}, I12: {sheet['I12'].value}, I13: {sheet['I13'].value}")
    print(f"I15: {sheet['I15'].value}, I16: {sheet['I16'].value}, I17: {sheet['I17'].value}")
