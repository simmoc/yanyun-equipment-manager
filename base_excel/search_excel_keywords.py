import openpyxl
import os

files = [
    "牵丝玉100级竞速轴属性毕业率计算器5.xlsx",
    "破竹风100级竞速轴属性毕业率计算器5.-1 copy.xlsx"
]

base_path = "/Users/simmoc/project/yanyun-equipment-manager/base_excel"

for filename in files:
    path = os.path.join(base_path, filename)
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"\nFile: {filename}")
    sheet = wb['期望']
    # Search for "ADPS" or "毕业率"
    for r in range(1, 100):
        for c in range(1, 30):
            val = sheet.cell(row=r, column=c).value
            if val and isinstance(val, str) and ("ADPS" in val or "毕业率" in val or "秒伤" in val):
                print(f"[{r},{c}]: {val} = {sheet.cell(row=r, column=c+1).value if c+1 < 30 else 'N/A'}")
